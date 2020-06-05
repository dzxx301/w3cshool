# !/usr/bin/python
# -*- coding:utf-8 -*-

import openpyxl

wb = openpyxl.load_workbook("E:/time.xlsx")   # 返回workbook数据类型
#ws = wb.get_sheet_by_name('原始数据')    # 获取表的名字
ws = wb['原始数据']      # 最新用法代替上面一行
# 2###########################
print(ws)
key_dat = ['2A','5A','3A','1A','11A','12A']
#  读取第一行表头的值
row_0 = [item.value for item in list(ws.rows)[0]]
print(row_0)

ws2 = wb.create_sheet(index=1,title="结果")    # 新建结果表单
ws2.append(row_0)    #  写入第1行表头的值

for x in key_dat:   # 根据关键数据进行循环
    for i in range(1,ws.max_row+1):     #  for i in range(1,mx_row+1):
        cell_value = ws.cell(row=i, column=2).value  # 获取第2列数值
        if cell_value == x:
            row_x =[item.value for item in list(ws.rows)[i-1]]
            print('第%d行:' % i,row_x)
            ws2.append(row_x)
        # else:
        #         #     print('表中没有',x)
# 1####################################
wb.save('E:/time.xlsx')
print("保存成功")


